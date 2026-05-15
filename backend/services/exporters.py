import csv
import io

EXPORT_COLUMNS = [
    "invoice_id",
    "original_name",
    "filename",
    "mime_type",
    "file_size",
    "uploaded_at",
    "status",
    "invoice_number",
    "invoice_date",
    "due_date",
    "vendor",
    "bill_to",
    "subtotal",
    "tax",
    "total",
    "extracted_at",
    "line_item_count",
    "line_items",
    "review_decision",
    "review_reason",
    "review_confidence",
]

EXPORT_VALIDATION_COLUMNS = [
    "is_valid",
    "date_format_ok",
    "amount_consistency_ok",
    "gst_calculation_ok",
    "total_match_ok",
]


def fetch_export_rows(db, organization_id: int, invoice_id=None):
    query = (
        "SELECT i.id AS invoice_id, i.original_name, i.filename, i.mime_type, i.file_size, i.uploaded_at, i.status, "
        "f.invoice_number, f.invoice_date, f.due_date, f.vendor, f.bill_to, f.subtotal, f.tax, f.total, f.extracted_at, "
        "COALESCE(li.line_item_count, 0) AS line_item_count, li.line_items, "
        "d.decision AS review_decision, d.reason AS review_reason, d.confidence_score AS review_confidence, "
        "v.is_valid, v.date_format_ok, v.amount_consistency_ok, "
        "v.gst_calculation_ok, v.total_match_ok "
        "FROM invoices i "
        "LEFT JOIN invoice_fields f ON i.id = f.invoice_id "
        "LEFT JOIN invoice_decisions d ON i.id = d.invoice_id "
        "LEFT JOIN ("
        "   SELECT vv.invoice_id, vv.is_valid, vv.date_format_ok, vv.amount_consistency_ok, vv.gst_calculation_ok, vv.total_match_ok "
        "   FROM invoice_validation vv "
        "   JOIN ("
        "       SELECT invoice_id, MAX(id) AS max_id "
        "       FROM invoice_validation "
        "       GROUP BY invoice_id"
        "   ) latest_v ON latest_v.max_id = vv.id"
        ") v ON i.id = v.invoice_id "
        "LEFT JOIN ("
        "   SELECT invoice_id, COUNT(*) AS line_item_count, "
        "          GROUP_CONCAT("
        "              CONCAT("
        "                  description, "
        "                  ' (qty=', COALESCE(CAST(quantity AS CHAR), ''), "
        "                  ', unit=', COALESCE(CAST(unit_price AS CHAR), ''), "
        "                  ', amount=', COALESCE(CAST(amount AS CHAR), ''), ')'"
        "              ) "
        "              SEPARATOR ' | '"
        "          ) AS line_items "
        "   FROM invoice_line_items "
        "   GROUP BY invoice_id"
        ") li ON i.id = li.invoice_id "
    )
    params = [organization_id]
    query += "WHERE i.organization_id = %s "
    if invoice_id:
        query += "AND i.id = %s "
        params.append(invoice_id)
    query += "ORDER BY i.uploaded_at DESC"
    with db.cursor() as cur:
        cur.execute(query, tuple(params))
        return cur.fetchall()


def build_csv(rows):
    output = io.StringIO()
    all_cols = EXPORT_COLUMNS + EXPORT_VALIDATION_COLUMNS
    writer = csv.DictWriter(output, fieldnames=all_cols, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow(row)
    return output.getvalue()


def build_excel(rows):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"

    all_cols = EXPORT_COLUMNS + EXPORT_VALIDATION_COLUMNS
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2E3440", end_color="2E3440", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, col_name in enumerate(all_cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name.replace("_", " ").title())
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, col_name in enumerate(all_cols, 1):
            value = row_data.get(col_name)
            if hasattr(value, "as_tuple"):
                value = float(value)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

    for col_idx, col_name in enumerate(all_cols, 1):
        max_len = max(len(col_name), 12)
        for row_data in rows:
            cell_val = str(row_data.get(col_name, "") or "")
            max_len = max(max_len, len(cell_val))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(
            max_len + 4,
            40,
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
